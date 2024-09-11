from PIL import Image, ImageFilter
import os
import time
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook, load_workbook

def redimensionar_imagem(imagem, tamanho=(800, 600)):
    return imagem.resize(tamanho)

def aplicar_filtros(imagem):
    imagem = imagem.filter(ImageFilter.EDGE_ENHANCE)
    imagem = imagem.convert('L')  
    return imagem

def processar_imagem_concorrente(arquivo, diretorio_origem, diretorio_destino):
    if arquivo.endswith(('jpeg', 'jpg', 'png')):
        caminho = os.path.join(diretorio_origem, arquivo)
        imagem = Image.open(caminho)
        
        imagem = redimensionar_imagem(imagem)
        imagem = aplicar_filtros(imagem)
        
        imagem.save(os.path.join(diretorio_destino, arquivo))

def processar_imagens_concorrente(diretorio_origem, diretorio_destino, num_threads, xlsx_arquivo):
    if not os.path.exists(diretorio_destino):
        os.makedirs(diretorio_destino)

    if os.path.exists(xlsx_arquivo):
        planilha = load_workbook(xlsx_arquivo)
    else:
        planilha = Workbook()

    aba_nome = f"Concorrente ({num_threads})"
    if aba_nome in planilha.sheetnames:
        aba_concorrente = planilha[aba_nome]
    else:
        aba_concorrente = planilha.create_sheet(title=aba_nome)

    inicio = time.time()

    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        arquivos = [f for f in os.listdir(diretorio_origem) if f.endswith(('jpeg', 'jpg', 'png'))]
        for arquivo in arquivos:
            executor.submit(processar_imagem_concorrente, arquivo, diretorio_origem, diretorio_destino)

    fim = time.time()
    tempo_execucao = fim - inicio

    aba_concorrente.append([f"Tempo total de execução (concorrente, {num_threads} threads)", f"{tempo_execucao:.2f} segundos"])

    planilha.save(xlsx_arquivo)
    print(f"Tempo total de execução (concorrente, {num_threads} threads): {tempo_execucao:.2f} segundos")

for threads in [2, 4, 8]:
    processar_imagens_concorrente('imagens_originais', f'imagens_processadas_concorrente_{threads}', threads, 'resultados.xlsx')
