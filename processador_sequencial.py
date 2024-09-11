from PIL import Image, ImageFilter
import os
import time
from openpyxl import Workbook, load_workbook

def redimensionar_imagem(imagem, tamanho=(800, 600)):
    return imagem.resize(tamanho)

def aplicar_filtros(imagem):
    imagem = imagem.filter(ImageFilter.EDGE_ENHANCE)
    imagem = imagem.convert('L')  
    return imagem

def processar_imagens_sequencial(diretorio_origem, diretorio_destino, xlsx_arquivo):
    if not os.path.exists(diretorio_destino):
        os.makedirs(diretorio_destino)

    # Criar ou abrir arquivo XLSX
    if os.path.exists(xlsx_arquivo):
        planilha = load_workbook(xlsx_arquivo)
    else:
        planilha = Workbook()

    # Selecionar ou criar aba para execução sequencial
    if "Execução Sequencial" in planilha.sheetnames:
        aba_sequencial = planilha["Execução Sequencial"]
    else:
        aba_sequencial = planilha.create_sheet(title="Execução Sequencial")

    inicio = time.time()
    
    for arquivo in os.listdir(diretorio_origem):
        if arquivo.endswith(('jpeg', 'jpg', 'png')):
            caminho = os.path.join(diretorio_origem, arquivo)
            imagem = Image.open(caminho)
            
            imagem = redimensionar_imagem(imagem)
            imagem = aplicar_filtros(imagem)
            
            imagem.save(os.path.join(diretorio_destino, arquivo))
    
    fim = time.time()
    tempo_execucao = fim - inicio

    # Adicionar dados à aba de execução sequencial
    aba_sequencial.append(["Tempo total de execução (sequencial)", f"{tempo_execucao:.2f} segundos"])

    planilha.save(xlsx_arquivo)
    print(f"Tempo total de execução (sequencial): {tempo_execucao:.2f} segundos")

# Executar o processamento sequencial e salvar dados
processar_imagens_sequencial('imagens_originais', 'imagens_processadas_sequencial', 'resultados.xlsx')
